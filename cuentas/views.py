"""
ViewSets avanzados para el módulo de cuentas.
Incluye filtros dinámicos, paginación, ordenamiento y optimizaciones de consultas.
También incluye vistas web para la interfaz de usuario.
"""
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Prefetch, Q
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView
from django.contrib import messages
from django.urls import reverse_lazy
from django.utils import timezone
from django.http import JsonResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io

from .models import (
    pais, provincia, localidad, tipo_documento, situacionIva, tipo_cuenta,
    cuenta, contacto_cuenta, direccion
)
from .serializers import (
    PaisSerializer, ProvinciaSerializer, LocalidadSerializer,
    TipoDocumentoSerializer, SituacionIvaSerializer, TipoCuentaSerializer,
    CuentaListSerializer, CuentaDetailSerializer, CuentaCreateUpdateSerializer,
    ContactoCuentaSerializer, DireccionSerializer
)
from .filters import CuentaFilter, ContactoCuentaFilter, DireccionFilter


class BaseViewSetMixin:
    """Mixin base con configuraciones comunes para todos los ViewSets."""
    
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    
    def get_queryset(self):
        """Optimizar queryset base con select_related."""
        queryset = super().get_queryset()
        
        # Aplicar optimizaciones comunes
        if hasattr(self, 'select_related_fields'):
            queryset = queryset.select_related(*self.select_related_fields)
        
        if hasattr(self, 'prefetch_related_fields'):
            queryset = queryset.prefetch_related(*self.prefetch_related_fields)
        
        return queryset


class PaisViewSet(BaseViewSetMixin, viewsets.ReadOnlyModelViewSet):
    """ViewSet para países (solo lectura)."""
    
    queryset = pais.objects.all().order_by('nombre')
    serializer_class = PaisSerializer
    search_fields = ['nombre']
    ordering_fields = ['id_pais', 'nombre']
    ordering = ['nombre']


class ProvinciaViewSet(BaseViewSetMixin, viewsets.ReadOnlyModelViewSet):
    """ViewSet para provincias con información del país."""
    
    queryset = provincia.objects.select_related('pais_idpais').order_by('nombre_provincia')
    serializer_class = ProvinciaSerializer
    search_fields = ['nombre_provincia', 'codigo_provincia']
    ordering_fields = ['id_provincia', 'nombre_provincia', 'codigo_provincia']
    ordering = ['nombre_provincia']
    select_related_fields = ['pais_idpais']
    
    @action(detail=False, methods=['get'])
    def por_pais(self, request):
        """Obtener provincias filtradas por país."""
        pais_id = request.query_params.get('pais_id')
        if not pais_id:
            return Response(
                {'error': 'Parámetro pais_id es requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        provincias = self.get_queryset().filter(pais_idpais=pais_id)
        serializer = self.get_serializer(provincias, many=True)
        return Response(serializer.data)


class LocalidadViewSet(BaseViewSetMixin, viewsets.ReadOnlyModelViewSet):
    """ViewSet para localidades con información jerárquica."""
    
    queryset = localidad.objects.select_related(
        'provincia_id_provincia__pais_idpais'
    ).order_by('nombre_localidad')
    serializer_class = LocalidadSerializer
    search_fields = ['nombre_localidad', 'cp_localidad']
    ordering_fields = ['id_localidad', 'nombre_localidad', 'cp_localidad']
    ordering = ['nombre_localidad']
    select_related_fields = ['provincia_id_provincia__pais_idpais']
    
    @action(detail=False, methods=['get'])
    def por_provincia(self, request):
        """Obtener localidades filtradas por provincia."""
        provincia_id = request.query_params.get('provincia_id')
        if not provincia_id:
            return Response(
                {'error': 'Parámetro provincia_id es requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        localidades = self.get_queryset().filter(provincia_id_provincia=provincia_id)
        serializer = self.get_serializer(localidades, many=True)
        return Response(serializer.data)


class TipoDocumentoViewSet(BaseViewSetMixin, viewsets.ReadOnlyModelViewSet):
    """ViewSet para tipos de documento."""
    
    queryset = tipo_documento.objects.all().order_by('descripcion')
    serializer_class = TipoDocumentoSerializer
    search_fields = ['descripcion']
    ordering_fields = ['idtipo_documento', 'descripcion', 'cod_afip']
    ordering = ['descripcion']


class SituacionIvaViewSet(BaseViewSetMixin, viewsets.ReadOnlyModelViewSet):
    """ViewSet para situaciones de IVA."""
    
    queryset = situacionIva.objects.all().order_by('descripcion')
    serializer_class = SituacionIvaSerializer
    search_fields = ['descripcion', 'reducida']
    ordering_fields = ['idsituacionIva', 'descripcion', 'codigo_afip']
    ordering = ['descripcion']


class TipoCuentaViewSet(BaseViewSetMixin, viewsets.ReadOnlyModelViewSet):
    """ViewSet para tipos de cuenta."""
    
    queryset = tipo_cuenta.objects.all().order_by('descripcion')
    serializer_class = TipoCuentaSerializer
    search_fields = ['descripcion']
    ordering_fields = ['id_tipo_cuenta', 'descripcion']
    ordering = ['descripcion']


class CuentaViewSet(BaseViewSetMixin, viewsets.ModelViewSet):
    """ViewSet principal para cuentas con funcionalidades avanzadas."""
    
    filterset_class = CuentaFilter
    search_fields = [
        'razon_social', 'nombre_fantasia', 'numero_documento', 
        'email_cuenta', 'telefono_cuenta', 'celular_cuenta'
    ]
    ordering_fields = [
        'id_cuenta', 'razon_social', 'nombre_fantasia', 
        'numero_documento', 'fecha_alta', 'activo'
    ]
    ordering = ['-id_cuenta']
    
    def get_queryset(self):
        """Queryset optimizado con selects y prefetches."""
        queryset = cuenta.objects.select_related(
            'tipo_documento_idtipo_documento',
            'situacionIva_idsituacionIva', 
            'tipo_cuenta_id_tipo_cuenta',
            'pais_id',
            'provincia_idprovincia',
            'localidad_idlocalidad'
        )
        
        # Para vistas de detalle, agregar relaciones anidadas
        if self.action == 'retrieve':
            queryset = queryset.prefetch_related(
                Prefetch(
                    'contactos',
                    queryset=contacto_cuenta.objects.filter(activo=True)
                ),
                'direcciones__pais_id',
                'direcciones__provincia_idprovincia',
                'direcciones__localidad_idlocalidad'
            )
        
        return queryset
    
    def get_serializer_class(self):
        """Seleccionar serializer según la acción."""
        if self.action == 'list':
            return CuentaListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return CuentaCreateUpdateSerializer
        else:
            return CuentaDetailSerializer
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtener estadísticas generales de cuentas."""
        queryset = self.filter_queryset(self.get_queryset())
        
        stats = {
            'total_cuentas': queryset.count(),
            'cuentas_activas': queryset.filter(activo=True).count(),
            'cuentas_inactivas': queryset.filter(activo=False).count(),
            'por_tipo': {}
        }
        
        # Estadísticas por tipo de cuenta
        tipos_stats = queryset.values(
            'tipo_cuenta_id_tipo_cuenta__descripcion'
        ).annotate(
            cantidad=Count('id_cuenta')
        ).order_by('-cantidad')
        
        for tipo_stat in tipos_stats:
            tipo_desc = tipo_stat['tipo_cuenta_id_tipo_cuenta__descripcion'] or 'Sin tipo'
            stats['por_tipo'][tipo_desc] = tipo_stat['cantidad']
        
        return Response(stats)
    
    @action(detail=True, methods=['post'])
    def activar(self, request, pk=None):
        """Activar una cuenta específica."""
        cuenta_obj = self.get_object()
        cuenta_obj.activo = True
        cuenta_obj.fecha_baja = None
        cuenta_obj.save()
        
        serializer = self.get_serializer(cuenta_obj)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def desactivar(self, request, pk=None):
        """Desactivar una cuenta específica."""
        from django.utils import timezone
        
        cuenta_obj = self.get_object()
        cuenta_obj.activo = False
        cuenta_obj.fecha_baja = timezone.now().date()
        cuenta_obj.save()
        
        serializer = self.get_serializer(cuenta_obj)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def contactos(self, request, pk=None):
        """Obtener contactos de una cuenta específica."""
        cuenta_obj = self.get_object()
        contactos = cuenta_obj.contactos.filter(activo=True)
        
        serializer = ContactoCuentaSerializer(contactos, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def direcciones(self, request, pk=None):
        """Obtener direcciones de una cuenta específica."""
        cuenta_obj = self.get_object()
        direcciones = cuenta_obj.direcciones.all()
        
        serializer = DireccionSerializer(direcciones, many=True)
        return Response(serializer.data)


class ContactoCuentaViewSet(BaseViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para contactos de cuenta."""
    
    serializer_class = ContactoCuentaSerializer
    filterset_class = ContactoCuentaFilter
    search_fields = ['nombre', 'cargo', 'email', 'telefono', 'celular']
    ordering_fields = ['id_contacto', 'nombre', 'cargo']
    ordering = ['-id_contacto']
    
    def get_queryset(self):
        """Queryset optimizado con select_related."""
        return contacto_cuenta.objects.select_related('cuenta_id')
    
    @action(detail=False, methods=['get'])
    def activos(self, request):
        """Obtener solo contactos activos."""
        queryset = self.filter_queryset(self.get_queryset()).filter(activo=True)
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class DireccionViewSet(BaseViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para direcciones."""
    
    serializer_class = DireccionSerializer
    filterset_class = DireccionFilter
    search_fields = ['etiqueta', 'calle', 'numero']
    ordering_fields = ['id_direccion', 'etiqueta', 'calle']
    ordering = ['-id_direccion']
    
    def get_queryset(self):
        """Queryset optimizado con select_related."""
        return direccion.objects.select_related(
            'cuenta_id',
            'pais_id',
            'provincia_idprovincia',
            'localidad_idlocalidad'
        )
    
    @action(detail=False, methods=['get'])
    def principales(self, request):
        """Obtener solo direcciones principales."""
        queryset = self.filter_queryset(self.get_queryset()).filter(es_principal=True)
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# ==========================================
# VISTAS WEB - INTERFAZ DE USUARIO
# ==========================================

class DashboardView(LoginRequiredMixin, TemplateView):
    """Vista del dashboard principal."""
    
    template_name = 'cuentas/dashboard.html'
    login_url = 'main:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas generales
        total_cuentas = cuenta.objects.count()
        cuentas_activas = cuenta.objects.filter(activo=True).count()
        
        # Cuentas del mes actual
        primer_dia_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        cuentas_mes_actual = cuenta.objects.filter(fecha_alta__gte=primer_dia_mes).count()
        
        # Provincias con cuentas
        provincias_con_cuentas = cuenta.objects.values('provincia_idprovincia__nombre_provincia').distinct().count()
        total_provincias = provincia.objects.count()
        
        # Última actualización
        ultima_cuenta = cuenta.objects.order_by('-fecha_alta').first()
        ultima_actualizacion = ultima_cuenta.fecha_alta if ultima_cuenta else timezone.now()
        
        # Cuentas recientes
        cuentas_recientes = cuenta.objects.select_related(
            'provincia_idprovincia', 'tipo_cuenta_id_tipo_cuenta', 'situacionIva_idsituacionIva'
        ).order_by('-fecha_alta')[:10]
        
        # Agregar nombre de provincia a cada cuenta
        for c in cuentas_recientes:
            c.provincia_nombre = c.provincia_idprovincia.nombre_provincia if c.provincia_idprovincia else None
            c.nombre_cuenta = c.razon_social  # Mapear para compatibilidad con template
            c.cuit_dni = c.numero_documento  # Mapear para compatibilidad con template
            c.fecha_creacion = c.fecha_alta  # Mapear para compatibilidad con template
            c.email = c.email_cuenta  # Mapear para compatibilidad con template
            c.telefono = c.telefono_cuenta  # Mapear para compatibilidad con template
        
        context.update({
            'stats': {
                'total_cuentas': total_cuentas,
                'cuentas_activas': cuentas_activas,
                'cuentas_mes_actual': cuentas_mes_actual,
                'porcentaje_activas': round((cuentas_activas / total_cuentas * 100) if total_cuentas > 0 else 0, 1),
                'provincias_con_cuentas': provincias_con_cuentas,
                'total_provincias': total_provincias,
                'ultima_actualizacion': ultima_actualizacion,
            },
            'cuentas_recientes': cuentas_recientes,
        })
        
        return context


class ListaCuentasView(LoginRequiredMixin, ListView):
    """Vista de lista de cuentas con filtros y búsqueda."""
    
    model = cuenta
    template_name = 'cuentas/lista_cuentas.html'
    context_object_name = 'cuentas'
    paginate_by = 20
    login_url = 'main:login'
    
    def get_queryset(self):
        queryset = cuenta.objects.select_related(
            'provincia_idprovincia',
            'localidad_idlocalidad',
            'tipo_cuenta_id_tipo_cuenta',
            'situacionIva_idsituacionIva',
            'tipo_documento_idtipo_documento'
        ).order_by('-fecha_alta')
        
        # Filtros
        busqueda = self.request.GET.get('q')
        if busqueda:
            queryset = queryset.filter(
                Q(razon_social__icontains=busqueda) |
                Q(numero_documento__icontains=busqueda) |
                Q(email_cuenta__icontains=busqueda)
            )
        
        tipo_cuenta_filter = self.request.GET.get('tipo_cuenta')
        if tipo_cuenta_filter:
            queryset = queryset.filter(tipo_cuenta_id_tipo_cuenta=tipo_cuenta_filter)
        
        provincia_filter = self.request.GET.get('provincia')
        if provincia_filter:
            queryset = queryset.filter(provincia_idprovincia=provincia_filter)
            
        localidad_filter = self.request.GET.get('localidad')
        if localidad_filter:
            queryset = queryset.filter(localidad_idlocalidad=localidad_filter)
        
        activo_filter = self.request.GET.get('activo')
        if activo_filter:
            queryset = queryset.filter(activo=activo_filter == 'true')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Datos para filtros
        context['tipos_cuenta'] = tipo_cuenta.objects.all()
        context['provincias'] = provincia.objects.all().order_by('nombre_provincia')
        context['localidades'] = localidad.objects.all().order_by('nombre_localidad')
        
        # Mantener valores de filtros
        context['filtros'] = {
            'q': self.request.GET.get('q', ''),
            'tipo_cuenta': self.request.GET.get('tipo_cuenta', ''),
            'provincia': self.request.GET.get('provincia', ''),
            'localidad': self.request.GET.get('localidad', ''),
            'activo': self.request.GET.get('activo', ''),
        }
        
        return context


class DetalleCuentaView(LoginRequiredMixin, DetailView):
    """Vista de detalle de cuenta."""
    
    model = cuenta
    template_name = 'cuentas/detalle_cuenta.html'
    context_object_name = 'cuenta'
    login_url = 'main:login'
    
    def get_queryset(self):
        return cuenta.objects.select_related(
            'provincia_idprovincia',
            'tipo_cuenta_id_tipo_cuenta',
            'situacionIva_idsituacionIva',
            'tipo_documento_idtipo_documento'
        ).prefetch_related(
            'contactos',
            'direcciones__pais_id',
            'direcciones__provincia_idprovincia',
            'direcciones__localidad_idlocalidad'
        )


class CrearCuentaView(LoginRequiredMixin, CreateView):
    """Vista para crear nueva cuenta."""
    
    model = cuenta
    template_name = 'cuentas/crear_cuenta.html'
    fields = [
        'razon_social', 'numero_documento', 'tipo_documento_idtipo_documento', 
        'tipo_cuenta_id_tipo_cuenta', 'situacionIva_idsituacionIva', 
        'pais_id', 'provincia_idprovincia', 'localidad_idlocalidad', 'direccion_cuenta',
        'telefono_cuenta', 'email_cuenta', 'activo'
    ]
    success_url = reverse_lazy('cuentas:lista_cuentas')
    login_url = 'main:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipos_documento'] = tipo_documento.objects.all()
        context['tipos_cuenta'] = tipo_cuenta.objects.all()
        context['situaciones_iva'] = situacionIva.objects.all()
        context['paises'] = pais.objects.all().order_by('nombre')
        context['provincias'] = provincia.objects.all().order_by('nombre_provincia')
        context['localidades'] = localidad.objects.all().order_by('nombre_localidad')
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Cuenta "{form.instance.razon_social}" creada exitosamente.')
        return super().form_valid(form)


class EditarCuentaView(LoginRequiredMixin, UpdateView):
    """Vista para editar cuenta existente."""
    
    model = cuenta
    template_name = 'cuentas/editar_cuenta.html'
    fields = [
        'razon_social', 'nombre_fantasia', 'numero_documento', 'tipo_documento_idtipo_documento', 
        'tipo_cuenta_id_tipo_cuenta', 'situacionIva_idsituacionIva', 
        'pais_id', 'provincia_idprovincia', 'localidad_idlocalidad', 'direccion_cuenta',
        'telefono_cuenta', 'email_cuenta', 'activo'
    ]
    success_url = reverse_lazy('cuentas:lista_cuentas')
    login_url = 'main:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipos_documento'] = tipo_documento.objects.all()
        context['tipos_cuenta'] = tipo_cuenta.objects.all()
        context['situaciones_iva'] = situacionIva.objects.all()
        context['paises'] = pais.objects.all().order_by('nombre')
        context['provincias'] = provincia.objects.all().order_by('nombre_provincia')
        context['localidades'] = localidad.objects.all().order_by('nombre_localidad')
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Cuenta "{form.instance.razon_social}" actualizada exitosamente.')
        return super().form_valid(form)


@login_required
def eliminar_cuenta_view(request, pk):
    """Vista para eliminar cuenta (AJAX)."""
    if request.method == 'POST':
        try:
            cuenta_obj = get_object_or_404(cuenta, id_cuenta=pk)
            nombre = cuenta_obj.razon_social
            cuenta_obj.delete()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Cuenta "{nombre}" eliminada exitosamente.'
                })
            else:
                messages.success(request, f'Cuenta "{nombre}" eliminada exitosamente.')
                return redirect('cuentas:lista_cuentas')
        
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Error al eliminar la cuenta.'
                })
            else:
                messages.error(request, 'Error al eliminar la cuenta.')
                return redirect('cuentas:lista_cuentas')
    
    return redirect('cuentas:lista_cuentas')


@login_required
def toggle_cuenta_activo_view(request, pk):
    """Vista para activar/desactivar cuenta (AJAX)."""
    if request.method == 'POST':
        try:
            cuenta_obj = get_object_or_404(cuenta, id_cuenta=pk)
            cuenta_obj.activo = not cuenta_obj.activo
            cuenta_obj.save()
            
            estado = 'activada' if cuenta_obj.activo else 'desactivada'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'activo': cuenta_obj.activo,
                    'message': f'Cuenta "{cuenta_obj.razon_social}" {estado} exitosamente.'
                })
            else:
                messages.success(request, f'Cuenta "{cuenta_obj.razon_social}" {estado} exitosamente.')
                return redirect('cuentas:lista_cuentas')
        
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Error al cambiar el estado de la cuenta.'
                })
            else:
                messages.error(request, 'Error al cambiar el estado de la cuenta.')
                return redirect('cuentas:lista_cuentas')
    
    return redirect('cuentas:lista_cuentas')


@login_required
def estadisticas_api_view(request):
    """API endpoint para estadísticas del dashboard (AJAX)."""
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Estadísticas en tiempo real
        total_cuentas = cuenta.objects.count()
        cuentas_activas = cuenta.objects.filter(activo=True).count()
        
        # Últimas 24 horas
        hace_24h = timezone.now() - timedelta(hours=24)
        cuentas_recientes_24h = cuenta.objects.filter(fecha_alta__gte=hace_24h).count()
        
        data = {
            'total_cuentas': total_cuentas,
            'cuentas_activas': cuentas_activas,
            'cuentas_inactivas': total_cuentas - cuentas_activas,
            'cuentas_recientes_24h': cuentas_recientes_24h,
            'porcentaje_activas': round((cuentas_activas / total_cuentas * 100) if total_cuentas > 0 else 0, 1),
            'ultima_actualizacion': timezone.now().isoformat(),
        }
        
        return JsonResponse(data)
    
    return JsonResponse({'error': 'Solo peticiones AJAX'}, status=400)


# ==========================================
# VISTAS AJAX PARA FILTROS EN CASCADA
# ==========================================

def ajax_provincias(request):
    """Vista AJAX para cargar provincias por país."""
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET:
        pais_id = request.GET.get('pais_id')
        
        if pais_id:
            provincias = provincia.objects.filter(pais_idpais_id=pais_id).order_by('nombre_provincia')
            data = {
                'provincias': [
                    {
                        'id': p.id_provincia,
                        'nombre': p.nombre_provincia
                    }
                    for p in provincias
                ]
            }
        else:
            data = {'provincias': []}
        
        return JsonResponse(data)
    
    return JsonResponse({'error': 'Solo peticiones AJAX'}, status=400)


def ajax_localidades(request):
    """Vista AJAX para cargar localidades por provincia."""
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET:
        provincia_id = request.GET.get('provincia_id')
        
        if provincia_id:
            localidades = localidad.objects.filter(provincia_id_provincia_id=provincia_id).order_by('nombre_localidad')
            data = {
                'localidades': [
                    {
                        'id': l.id_localidad,
                        'nombre': l.nombre_localidad,
                        'cp': l.cp_localidad
                    }
                    for l in localidades
                ]
            }
        else:
            data = {'localidades': []}
        
        return JsonResponse(data)
    
    return JsonResponse({'error': 'Solo peticiones AJAX'}, status=400)


@login_required
def exportar_cuentas_pdf(request):
    """Exportar listado de cuentas a PDF."""
    # Obtener todas las cuentas
    cuentas = cuenta.objects.select_related(
        'tipo_cuenta_id_tipo_cuenta',
        'tipo_documento_idtipo_documento',
        'situacionIva_idsituacionIva'
    ).all()
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="listado_cuentas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        textColor=colors.HexColor('#2c3e50'),
        alignment=1  # Centrado
    )
    
    # Título
    story.append(Paragraph("Listado de Cuentas - AGL SRL", title_style))
    story.append(Spacer(1, 12))
    
    # Preparar datos para la tabla
    data = [['ID', 'Razón Social', 'Tipo Documento', 'Número Doc.', 'Email']]
    
    for c in cuentas:
        data.append([
            str(c.id_cuenta),
            c.razon_social or '',
            c.tipo_documento_idtipo_documento.descripcion if c.tipo_documento_idtipo_documento else '',
            c.numero_documento or '',
            c.email_cuenta or ''
        ])
    
    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(table)
    
    # Información adicional
    story.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey
    )
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", footer_style))
    story.append(Paragraph(f"Total de cuentas: {len(cuentas)}", footer_style))
    
    # Generar PDF
    doc.build(story)
    
    return response


@login_required
def exportar_cuentas_excel(request):
    """Exportar listado de cuentas a Excel."""
    # Obtener todas las cuentas
    cuentas = cuenta.objects.select_related(
        'tipo_cuenta_id_tipo_cuenta',
        'tipo_documento_idtipo_documento',
        'situacionIva_idsituacionIva'
    ).all()
    
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listado de Cuentas"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Razón Social', 'Nombre Fantasía', 'Tipo Documento', 'Número Documento',
        'Dirección', 'Teléfono', 'Celular', 'Email', 'Tipo Cuenta',
        'Situación IVA', 'Activo', 'Fecha Alta'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row_num, c in enumerate(cuentas, 2):
        ws.cell(row=row_num, column=1, value=c.id_cuenta)
        ws.cell(row=row_num, column=2, value=c.razon_social or '')
        ws.cell(row=row_num, column=3, value=c.nombre_fantasia or '')
        ws.cell(row=row_num, column=4, value=c.tipo_documento_idtipo_documento.descripcion if c.tipo_documento_idtipo_documento else '')
        ws.cell(row=row_num, column=5, value=c.numero_documento or '')
        ws.cell(row=row_num, column=6, value=c.direccion_cuenta or '')
        ws.cell(row=row_num, column=7, value=c.telefono_cuenta or '')
        ws.cell(row=row_num, column=8, value=c.celular_cuenta or '')
        ws.cell(row=row_num, column=9, value=c.email_cuenta or '')
        ws.cell(row=row_num, column=10, value=c.tipo_cuenta_id_tipo_cuenta.descripcion if c.tipo_cuenta_id_tipo_cuenta else '')
        ws.cell(row=row_num, column=11, value=c.situacionIva_idsituacionIva.descripcion if c.situacionIva_idsituacionIva else '')
        ws.cell(row=row_num, column=12, value='Sí' if c.activo else 'No')
        ws.cell(row=row_num, column=13, value=c.fecha_alta.strftime('%d/%m/%Y') if c.fecha_alta else '')
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="listado_cuentas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Guardar el workbook en la respuesta
    wb.save(response)
    
    return response
